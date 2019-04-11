# -*- coding: utf-8 -*-
"""
Created on Wed Apr  3 10:22:25 2019

@author: jamesbriggs
"""
import os
import json
import time
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.support.ui import Select


def format_df(df):
    """
    This function will format the dataframe, eg converting Launch Date values
    into datetime objects and using regular expressions to  fix weird currency
    formats (such as [41.5m] in the 'Face Value' column).
    """
    # setting up vectorised date conversion function
    def date_convert(x):
        return datetime.strptime(x, '%d %b %Y')

    v_date_convert = np.vectorize(date_convert)

    df['Launch Date'] = v_date_convert(df['Launch Date'])

    # !!! will need to do face value thing later as also needs currency conversion

    return df


class webscrape:
    
    def __init__(self):
        """
        Here will find when last web scrape was conducted.
        """
        detail_loc = r'properties\details.json'
        if not os.path.isfile(detail_loc):
            # if no details JSON object exists, create new using earliest
            # start date
            self.details = {
                    'previous': '01 Jan 2010'
                    }
            # saving to details.json
            with open(detail_loc, 'w') as fp:
                json.dump(self.details, fp)

        else:
            # if details JSON object does exist, just load
            with open(detail_loc, 'r') as fp:
                self.details = json.load(fp)

    def download(self, sleeptime):
        """
        Here we will pull all data from the previous web scrape to the current
        date. The sleeptime argument controls the time that the program will
        wait in order to allow webpages to load, adjust to suit internet
        speeds.
        """

        # getting the current date in the required format
        # eg = dd MMM yyyy (ex 01 Mar 2019)
        current_date = datetime.strftime(datetime.today(), '%d %b %Y')
        
        # and for the previous week
        prior_date = self.details['previous']

        # show user what date range we will be using
        print('Pulling data from {} to {}.'.format(prior_date, current_date))
        
        # setting up Selenium web driver
        driver = webdriver.Chrome('drivers/chromedriver.exe')
        
        # opening conceptABS webpage
        driver.get('https://www.conceptabs.co.uk/default.aspx')

        # get user and pass from secret folder
        with open('secret/user.txt', 'r') as fp:
            user = fp.read()

        with open('secret/pass.txt', 'r') as fp:
            pw = fp.read()

        """
        Logging in
        """
        # so we can track the number of login attempts and stop process if too high
        attempt = 0
        # putting into a while statement to repeat until we are allowed in
        while True:
            attempt += 1
            print('Login attempt {}'.format(attempt))
        
            # giving page a moment to load
            time.sleep(sleeptime)
            # username
            driver.find_element_by_name('ctl00$CPH1$LoginView1$Login1$UserName').send_keys(user)
        
            # password
            driver.find_element_by_name('ctl00$CPH1$LoginView1$Login1$Password').send_keys(pw)
        
            # press login button
            driver.find_element_by_name('ctl00$CPH1$LoginView1$Login1$LoginImageButton').click()
        
            """
            Navigating to CDOs
            """
            # give the page a moment to load
            time.sleep(sleeptime)
            # at this point we may need to login again, so will use try-except statement
            try:
                # click the search database and graphing option
                driver.find_element_by_name('ctl00$CPH1$Heads$PicLink2').click()
                # if successful, break from the while-loop
                break
            except:
                # if we need to login again...
                driver.find_element_by_name('ctl00$CPH1$LoginView1$btnLoginRed').click()
                # stop process if attempt number too high
                if attempt > 9:
                    driver.close()
                    raise ValueError('Too many refused login attempts. Driver closed.')
                
        # give page time to load
        time.sleep(sleeptime)
        
        # click the Date tab
        driver.find_element_by_id('__tab_ctl00_CPH1_TC3_tpDate').click()
        
        time.sleep(sleeptime)
            
        # select first box (select current.date - 1 week)
        driver.find_element_by_name('ctl00$CPH1$TC3$tpDate$Date1').send_keys(prior_date)
        
        # select second box (select current.date)
        driver.find_element_by_id('ctl00_CPH1_TC3_tpDate_Date2').click()
        time.sleep(sleeptime)
        driver.find_element_by_id('ctl00_CPH1_TC3_tpDate_Date2').send_keys(current_date)
        
        attempt = 0
        
        while True:
            attempt += 1
            print('Pulling data attempt {}'.format(attempt))
            time.sleep(sleeptime)
        
            # select + add filter
            driver.find_element_by_name('ctl00$CPH1$TC3$tpDate$btnDateAdd').click()
            time.sleep(sleeptime)
            
            # press search
            driver.find_element_by_name('ctl00$CPH1$btnSearch').click()
            time.sleep(sleeptime)
            
            # let page load
            time.sleep(sleeptime)
            
            # get list of downloads folder before download
            pre_dl = os.listdir(r'C:\Users\jamesbriggs\Downloads')
            
            try:
                # select tranches
                driver.find_element_by_name('ctl00$CPH1$btnExportTranches').click()
                break
            except:
                if attempt > 9:
                    driver.close()
                    raise ValueError('Too many failed data pull attempts. Driver closed.')
                pass
        
        # wait 5 seconds to allow download
        time.sleep(5)
        
        # get list of downloads folder before download
        post_dl = os.listdir(r'C:\Users\jamesbriggs\Downloads')
        
        # find the difference between pre and post DL lists
        new_file = list(set(post_dl) - set(pre_dl))
        
        if len(new_file) == 1:
            print('New files:\n{}'.format(new_file[0]))
            
            # moving file
            os.rename(r'C:\Users\jamesbriggs\Downloads\{}'.format(new_file[0]),
                      r'C:\Users\jamesbriggs\Documents\Web Scraping\ConceptABS\data\{}'.format(new_file[0]))
            
        else:
            raise IOError('No new files downloaded to Downloads directory.')
        
        # finish
        driver.close()

        # return the filename of document downloaded
        return new_file[0]

    def scrape_all(self):
        """
        Here will scrape all data from the set starting date to now. This
        will overwrite the data.csv file.
        """
        pass

    """
    =======================================================================

    """


class data:
    
    def __init__(self):
        """
        Here we will check for the the full dataframe located in the data dir
        with the filename 'data.csv'. If the data file is not present we will
        initialise an empty dataframe with the correct headers ready to append
        to.
        """
        # set the target location
        target = r'data\data.csv'

        # check if the data.csv file exists, if not, initialise a new file
        if not os.path.isfile(target):
            # created a new dataframe object
            self.current = pd.DataFrame()
        else:
            # load the current data
            self.current = pd.read_csv(target)
        

    def add(self, filename):
        """
        Clean the given filename CABS search results file and append to the
        combined datatable in the data\full directory.
        """
        # set the target file
        target = r'data\{}'.format(filename)
        # first we will remove the top 5 rows in the CABS file so we are left
        # with only headings and data
        print(r'Loading .\{}'.format(target))
        # loading the workbook
        wb = load_workbook(filename=target)
        # selecting the worksheet
        ws = wb.active
        # before deleting the top 5 rows, we must check that they are empty
        formatted = False
        for i in range(1, 6):
            print(i)
            if ws['C{}'.format(i)].value != None:
                print('Sheet already formatted.')
                formatted = True
                break

        # if the sheet is not formatted, we will format
        if not formatted:
            print('Formatting sheet.')
            # deleting top 5 rows
            ws.delete_rows(1, 5)
            # saving modified workbook
            wb.save(target)

        # now we can pull the formatted data into Pandas
        new = pd.read_excel(target)

        # append the new data to the self.current dataframe
        self.current = self.current.append(new, ignore_index=True)

        # !!! remove duplicates - confirm with Laura that this is okay
        print('line 238: check with Laura this is okay')
        self.current.drop_duplicates(inplace=True)

        # now save the new data object
        self.current.to_csv(r'data\data.csv')

        # update the user on progress
        print('\'{}\' data appended to data file'.format(filename))

        # now delete the uneeded CABS Search Results excel sheet
        os.remove(target)

        # update the properties/details.json file with the most recent pull date
        # first we specify the target file...
        detail_loc = r'properties\details.json'
        # now we load the current detail.json
        with open(detail_loc, 'r') as fp:
            detail = json.load(fp)
        # update the 'previous' property with the current datetime (formatted)
        detail['previous'] = datetime.strftime(datetime.now(), '%d %b %Y')
        # finally, save the updated date to details.json
        with open(detail_loc, 'w') as fp:
             json.dump(detail, fp)
    
    """
    =======================================================================

    """


class visualise:
    """
    Visualisation class used to generate the figures and dashboards using the
    conceptABS data.
    """
    # YOU SHOULD LOOK INTO https://matplotlib.org/gallery/misc/multipage_pdf.html#sphx-glr-gallery-misc-multipage-pdf-py FOR REPORT CREATION (LOOKS PROMISING - USING LaTeX ALSO)
    def __init__(self):
        """
        Here we import the conceptABS data and the most recent data pull date.
        We also check if the outputs directory exists, if not, we create it.
        """
        # pulling in the conceptABS data
        self.data = pd.read_csv(r'data\data.csv')
        # cleaning the data
        self.data = format_df(self.data)

        # pulling in the most recent pull-date
        with open(r'properties\details.json', 'r') as fp:
            detail = json.load(fp)

        self.pull_date = detail['previous']  # !!! maybe should update 'previous' to be 'pull-date'

        # check if output directory exists, if not, create it
        if not os.path.isdir(r'reports\vis'):
            os.mkdir(r'reports\vis')

    def html(self):
        """
        Here we will generate an excel-based dashboard using figures built in
        matplotlib and seaborn.
        """
        # YOU SHOULD CONSIDER CREATING A PURE HTML/CSS WEB-REPORT BY MAKING A PRESET HTML STRING AND POPULATING IT WITH THE RELEVANT INFO, THIS WILL ALSO INCLUDE THE FIGURES BUILT - COULD INCLUDE MULTIPLE PAGES ETC
        # WHICH COULD THEN AUTOMATICALLY COMPILE INTO A ZIP FILE READY TO BE SENT TO CLIENTS

        # getting a monthly date column (eg day = 01, time = 00:00:00)
        # setting up function
        def month_dt(x):
            # convert to string containing just month and year then back to
            # datetime object which will assume day=01 and time=00:00:00
            x = datetime.strftime(x, '%m-%Y')
            return datetime.strptime(x, '%m-%Y')

        # creating the monthly launch date column
        self.data['Monthly Launch Date'] = self.data['Launch Date'].apply(month_dt)

        # pulling the visuals.xlsx mappings file
        visuals_df = pd.read_excel(r'properties\visuals.xlsx')

        """
        Plotting functions
        """

        def time_series(x, y, ylabel, title):
            """
            Function for plotting a time-series of y against x (time).
            """
            # create plot
            sns.lineplot(x=x, y=y)

            # label settings
            plt.xlabel('Time')
            plt.ylabel(ylabel)

            # tighten layout (to remove unwanted whitespace)
            plt.tight_layout()

            # save plot
            plt.savefig(r'reports\vis\{}.png'.format(title))
            
        def barchart(x, y, xlabel, ylabel, title):
            """
            Function for plotting a simple barchart.
            """
            # create plot
            sns.barplot(x=x, y=y)

            # label settings
            plt.xlabel(xlabel)
            plt.ylabel(ylabel)
            plt.xticks(rotation=45)

            # tighten layout
            plt.tight_layout()

            # save plot
            plt.savefig(r'reports\vis\{}.png'.format(title))
            
        def plotter(data, row):
            """
            This function controls the mapping -> visualisation logic. It also
            initialises the figure and colour palette.
            """
            # basic figure setup
            plt.figure(figsize=(12,8))
            sns.set_style('darkgrid')
            # colour palette
            sns.set_palette(['#86BC25', '#000000', '#00A3E0',
                             '#012169', '#D0D0CE', '#2C5234',
                             '#43B02A', '#046A38', '#BBBCBC',
                             '#97999B', '#0097A9', '#62B5E5',
                             '#C4D600', '#0076A8', '#75787B',
                             '#53565A', '#E3E48D', '#009A44',
                             '#DDEFE8', '#9DD4CF', '#A0DCFF',
                             '#005587', '#041E42', '#A7A8AA',
                             '#6FC2B4', '#00ABAB', '#007680',
                             '#004F59', '#63666A'])

            # checking the required figure type
            if row['Type'] == 'time series':
                # plot time series plot
                time_series(self.data[row['x-axis']], self.data[row['y-axis']],
                            row['y Label'], row['Title'])
            elif row['Type'] == 'barchart':
                # check if we want to filter to top n categories (x-axis)
                if str(row['Top']) != 'nan':
                    top = int(row['Top'])
                    # try grouping x and summing y first
                    try:
                        data = data.groupby(by=row['x-axis']).sum()[[row['y-axis']]].reset_index()
                    except TypeError:
                        data = data.groupby(by=row['x-axis']).count()[[row['y-axis']]].reset_index()
                    # now filter so we only get top n number of rows
                    data = data.nlargest(top, row['y-axis'])
                    
                # now plot
                barchart(data[row['x-axis']], data[row['y-axis']],
                         row['x Label'], row['y Label'], row['Title'])

        """
        Building the HTML sheets.
        """
        #######################################
        ###             INDEX               ###
        #######################################

        # start with the top of the page (before any custom figures or text)
        index = '''<!DOCTYPE html>
<html lang="en">

<head>

    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1, shrink-to-fit=no">
    <meta name="description" content="">
    <meta name="author" content="">

    <title>Deloitte Valuations</title>

    <!-- Bootstrap core CSS -->
    <link href="template/bootstrap.min.css" rel="stylesheet">

</head>

<body>

    <!-- Navigation -->
    <nav class="navbar navbar-expand-lg navbar-dark bg-dark static-top">
        <div class="container">
            <a class="navbar-brand" href="#">CDO Valuations</a>
            <button class="navbar-toggler" type="button" data-toggle="collapse" data-target="#navbarResponsive" aria-controls="navbarResponsive" aria-expanded="false" aria-label="Toggle navigation">
                <span class="navbar-toggler-icon"></span>
            </button>
            <div class="collapse navbar-collapse" id="navbarResponsive">
                <ul class="navbar-nav ml-auto">
                    <li class="nav-item active">
                        <a class="nav-link" href="#">Summary
                            <span class="sr-only">(current)</span>
                        </a>
                    </li>
                    <li class="nav-item">
                        <a class="nav-link" href="https://www.youtube.com/watch?v=7ohbr90cKDc">Breakdown</a>
                    </li>
                    <li class="nav-item">
                        <a class="nav-link" href="portfolio.html">Portfolios</a>
                    </li>
                    <li class="nav-item">
                        <a class="nav-link" href="#">Data</a>
                    </li>
                </ul>
            </div>
        </div>
    </nav>

    <!-- Page Content -->
    <div class="container">
        <div class="row">
            <div class="col-lg-12 text-center">
                <h1 class="mt-5">Summary</h1>
                <p class="lead">An overview of CDO data.</p>
'''
        # now we will add summary figures - first we get summary rows only
        summary_df = visuals_df[visuals_df['Section'] == 'Summary']

        for i in range(len(summary_df)):
            # grabbing data row by row
            row = summary_df.iloc[i]

            # create figure
            plotter(self.data, row)

            # and now append this plot filepath to the html code
            index += '''
                    <h4>{}</h4>
                        <img src="{}"/>
                    <p>{}</p>
                    '''.format(row['Title'],
                    'vis/{}.png'.format(row['Title']),
                    row['Description'])

        # now we are done generating summary figures and their HTML, append
        # final part of index HTML
        index += '''
                <ul class="list-unstyled">
                    <li>Last Updated: {}</li>
                </ul>
            </div>
        </div>
    </div>

    <!-- Bootstrap core JavaScript -->
    <script src="template/jquery.min.js"></script>
    <script src="template/bootstrap.bundle.min.js"></script>

</body>

</html>
'''.format(self.pull_date)

        # now save the formatted html index document
        fp = open(r'reports\index.html', 'w')
        fp.write(index)
        fp.close()
        
        #######################################
        ###            PORTFOLIO            ###
        #######################################

        # start with the top of the page (before any custom figures or text)
        portfolio = '''<!DOCTYPE html>
<html lang="en">

<head>

    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1, shrink-to-fit=no">
    <meta name="description" content="">
    <meta name="author" content="">

    <title>Deloitte Valuations</title>

    <!-- Bootstrap core CSS -->
    <link href="template/bootstrap.min.css" rel="stylesheet">

</head>

<body>

    <!-- Navigation -->
    <nav class="navbar navbar-expand-lg navbar-dark bg-dark static-top">
        <div class="container">
            <a class="navbar-brand" href="#">CDO Valuations</a>
            <button class="navbar-toggler" type="button" data-toggle="collapse" data-target="#navbarResponsive" aria-controls="navbarResponsive" aria-expanded="false" aria-label="Toggle navigation">
                <span class="navbar-toggler-icon"></span>
            </button>
            <div class="collapse navbar-collapse" id="navbarResponsive">
                <ul class="navbar-nav ml-auto">
                    <li class="nav-item">
                        <a class="nav-link" href="index.html">Summary</a>
                    </li>
                    <li class="nav-item">
                        <a class="nav-link" href="https://www.youtube.com/watch?v=7ohbr90cKDc">Breakdown</a>
                    </li>
                    <li class="nav-item">
                        <a class="nav-link active" href="#">Portfolios
                            <span class="sr-only">(current)</span>
                        </a>
                    </li>
                    <li class="nav-item">
                        <a class="nav-link" href="#">Data</a>
                    </li>
                </ul>
            </div>
        </div>
    </nav>

    <!-- Page Content -->
    <div class="container">
        <div class="row">
            <div class="col-lg-12 text-center">
                <h1 class="mt-5">Portfolios</h1>
                <p class="lead">An overview of bank portfolios.</p>
'''
        # now we will add summary figures - first we get summary rows only
        portfolio_df = visuals_df[visuals_df['Section'] == 'Portfolio']

        for i in range(len(portfolio_df)):
            # grabbing data row by row
            row = portfolio_df.iloc[i]

            # create figure
            plotter(self.data, row)

            # and now append this plot filepath to the html code
            portfolio += '''
                    <h4>{}</h4>
                        <img src="{}"/>
                    <p>{}</p>
                    '''.format(row['Title'],
                    'vis/{}.png'.format(row['Title']),
                    row['Description'])

        # now we are done generating summary figures and their HTML, append
        # final part of index HTML
        portfolio += '''
                <ul class="list-unstyled">
                    <li>Last Updated: {}</li>
                </ul>
            </div>
        </div>
    </div>

    <!-- Bootstrap core JavaScript -->
    <script src="template/jquery.min.js"></script>
    <script src="template/bootstrap.bundle.min.js"></script>

</body>

</html>
'''.format(self.pull_date)

        # now save the formatted html index document
        fp = open(r'reports\portfolio.html', 'w')
        fp.write(portfolio)
        fp.close()
            

    
    def interactive(self):
        """
        Here we will initialise a simple web server to use for interactive
        d3.js visualisations.
        - don't get distracted by this part - do it last
        """
        pass

    """
    =======================================================================

    """
